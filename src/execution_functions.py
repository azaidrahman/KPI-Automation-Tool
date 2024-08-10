import pandas as pd
import os
from datetime import datetime
import numpy as np
from tkinter import messagebox
import traceback

from processing_functions import process_report, initialize_main_dataframe, start_date, end_date, process_budget, process_payroll, process_workbills, remove_empty_rows_and_columns

from loading_functions import read_job_list, read_budget, read_payroll, read_workbills

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def load_data(job_list_files, budget_files, payroll_files, workbills_files):
    job_list_df = read_job_list(job_list_files[0]) if job_list_files else None
    
    error_dict = {}
    
    budget_dfs = [read_budget(file, error_dict) for file in budget_files]
    budget_dfs = [df for df in budget_dfs if df is not None]
    
    payroll_dfs = [read_payroll(file, error_dict) for file in payroll_files]
    payroll_dfs = [df for df in payroll_dfs if df is not None]
    
    workbills_dfs = [read_workbills(file, error_dict) for file in workbills_files]
    workbills_dfs = [df for df in workbills_dfs if df is not None]
    
    return job_list_df, budget_dfs, payroll_dfs, workbills_dfs, error_dict

def process_data(job_list_df, budget_dfs, payroll_dfs, workbills_dfs):
    main_df = initialize_main_dataframe(job_list_df, start_date, end_date)
    
    for budget_df in budget_dfs:
        budget_data_df = process_budget(budget_df)
        main_df = process_report('budget', job_list_df, budget_data_df, main_df)
    
    for payroll_df in payroll_dfs:
        payroll_data_df = process_payroll(payroll_df)
        main_df = process_report('payroll', job_list_df, payroll_data_df, main_df)
    
    for workbills_df in workbills_dfs:
        workbills_data_df = process_workbills(workbills_df)
        main_df = process_report('workbills', job_list_df, workbills_data_df, main_df)
    
    return main_df

def apply_conditional_formatting(csm_sheet):
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Light gray
    
    # Iterate over the rows, starting from row 4 where data begins
    for row_idx in range(4, csm_sheet.max_row + 1, 2):  # Step by 2 for planned vs actual pairs
        for col_idx in range(4, csm_sheet.max_column + 1, 2):  # Start from column D (4) which should now be YTD
            planned_hours = csm_sheet.cell(row=row_idx, column=col_idx).value
            planned_cost = csm_sheet.cell(row=row_idx, column=col_idx+1).value
            actual_hours = csm_sheet.cell(row=row_idx+1, column=col_idx).value
            actual_cost = csm_sheet.cell(row=row_idx+1, column=col_idx+1).value
            
            # Function to check if a value is effectively empty
            def is_empty(value):
                return value is None or value == '' or (isinstance(value, float) and (np.isnan(value) or value == 0))
            
            # Check for partially filled pairs
            if is_empty(planned_hours) and not is_empty(actual_hours):
                csm_sheet.cell(row=row_idx+1, column=col_idx).fill = gray_fill
                csm_sheet.cell(row=row_idx+1, column=col_idx+1).fill = gray_fill
                continue
            elif not is_empty(planned_hours) and is_empty(actual_hours):
                csm_sheet.cell(row=row_idx, column=col_idx).fill = gray_fill
                csm_sheet.cell(row=row_idx, column=col_idx+1).fill = gray_fill
                continue
            
            # Skip completely empty pairs
            if is_empty(planned_hours) and is_empty(actual_hours):
                continue
            
            # Convert to float for comparison
            try:
                planned_hours = float(planned_hours)
                planned_cost = float(planned_cost)
                actual_hours = float(actual_hours)
                actual_cost = float(actual_cost)
            except (ValueError, TypeError):
                continue  # Skip if conversion to float fails
            
            if actual_hours < planned_hours:
                csm_sheet.cell(row=row_idx+1, column=col_idx).fill = green_fill
                csm_sheet.cell(row=row_idx+1, column=col_idx+1).fill = green_fill
            elif abs(planned_hours - actual_hours) <= 1:
                csm_sheet.cell(row=row_idx+1, column=col_idx).fill = yellow_fill
                csm_sheet.cell(row=row_idx+1, column=col_idx+1).fill = yellow_fill
            else:
                csm_sheet.cell(row=row_idx+1, column=col_idx).fill = red_fill
                csm_sheet.cell(row=row_idx+1, column=col_idx+1).fill = red_fill

def format_csm_sheet(csm_sheet, csm_name):
    csm_sheet['A2'] = csm_name
    csm_sheet['A2'].font = Font(size=16, bold=True)
    csm_sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in csm_sheet['A2:A2'][0]:
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    
    csm_sheet['C1'] = 'Bi-weekly period'
    
    for cell in csm_sheet[1][3:]:
        if isinstance(cell.value, datetime):
            cell.value = cell.value.date()
            cell.number_format = 'YYYY-MM-DD'
    
    csm_sheet['A1'] = "Back to Index"
    csm_sheet['A1'].hyperlink = "#Index!A1"
    csm_sheet['A1'].style = "Hyperlink"
    
    apply_conditional_formatting(csm_sheet)
    csm_sheet.freeze_panes = 'D3'
    
    for i in range(1, 4):
        column_letter = get_column_letter(i)
        max_length = max(len(str(cell.value)) for cell in csm_sheet[column_letter] if cell.value)
        csm_sheet.column_dimensions[column_letter].width = max_length
        

def write_excel(main_df, output_path):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_filename = os.path.basename(os.path.splitext(output_path)[0])
    output_dir = os.path.dirname(output_path)
    output_ext = ".xlsx"  # Changed to .xlsx
    counter = 1
    final_output_path = os.path.join(output_dir, f"KPI CSM {output_filename}_{timestamp}{output_ext}")
    
    while os.path.exists(final_output_path):
        final_output_path = os.path.join(output_dir, f"KPI CSM {output_filename}_{timestamp}_{counter}{output_ext}")
        counter += 1
    
    os.makedirs(output_dir, exist_ok=True)

    test_count = 0
    
    with pd.ExcelWriter(final_output_path, engine='openpyxl') as writer:
        if len(main_df) == 0:
            raise ValueError("No data to write to Excel.")
        
        csm_names = sorted(main_df.index.get_level_values('csm_name').unique())
        
        index_sheet = writer.book.create_sheet(title='Index')
        index_sheet['A1'] = 'CSM Name'
        index_sheet['B1'] = 'Link'
        
        sheets_written = 0
        for index, csm_name in enumerate(csm_names, start=2):
            csm_df = main_df.xs(csm_name, level='csm_name')
            
            if not csm_df.empty:
                csm_df = remove_empty_rows_and_columns(csm_df)
                
                if not csm_df.empty:
                    # Calculate YTD for Hours and Cost, excluding NaN values
                    ytd_hours = csm_df.xs('Hours', axis=1, level=1).apply(lambda x: x.sum() if x.notna().any() else pd.NA, axis=1)
                    ytd_cost = csm_df.xs('Cost', axis=1, level=1).apply(lambda x: x.sum() if x.notna().any() else pd.NA, axis=1)
                    
                    # Create YTD columns
                    ytd = pd.concat([ytd_hours, ytd_cost], axis=1, keys=['Hours', 'Cost'])
                    ytd.columns = pd.MultiIndex.from_product([['YTD'], ytd.columns])
                    
                    # Insert YTD columns after the 'type_of_KPI' column
                    result_df = pd.concat([csm_df.iloc[:, :0], ytd, csm_df.iloc[:, 0:]], axis=1)
                    # test_count += 1
                    # if test_count <= 1: print(result_df)  # Test

                    sheets_written += 1
                    result_df.to_excel(writer, sheet_name=csm_name)
                    
                    index_sheet[f'A{index}'] = csm_name
                    index_sheet[f'B{index}'].hyperlink = f"#'{csm_name}'!A1"
                    index_sheet[f'B{index}'].style = "Hyperlink"
                    index_sheet[f'B{index}'] = "Go to Sheet"
                    
                    csm_sheet = writer.sheets[csm_name]
                    format_csm_sheet(csm_sheet, csm_name)
                    
    if sheets_written == 0:
        raise ValueError("No valid data sheets created for the Excel file.")
    
    return final_output_path

def process_files(job_list_files, budget_files, payroll_files, workbills_files, output_path):

    try:
        job_list_df, budget_dfs, payroll_dfs, workbills_dfs, error_dict = load_data(job_list_files, budget_files, payroll_files, workbills_files)
        
        if error_dict:
            error_message = "The following files had issues with date parsing:\n\n"
            for file, error in error_dict.items():
                error_message += f"{file}: {error}\n"
            messagebox.showwarning("Date Parsing Issues", error_message)
        
        # Check if any required dataset is missing or empty
        if (job_list_df is None or job_list_df.empty or
            not budget_dfs or all(df.empty for df in budget_dfs) or
            not payroll_dfs or all(df.empty for df in payroll_dfs) or
            not workbills_dfs or all(df.empty for df in workbills_dfs)):
            raise ValueError("One or more required datasets are empty after processing.")
        
        main_df = process_data(job_list_df, budget_dfs, payroll_dfs, workbills_dfs)
        
        output_path = os.path.splitext(output_path)[0] + '.xlsm'
        
        final_output_path = write_excel(main_df, output_path)
        print(f"CSM KPI SHEET generated at: {final_output_path}")
        return final_output_path
    except Exception as e:
        error_message = f"Error while processing files: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_message)
        messagebox.showerror("Error", error_message)
        with open("../error_log.txt", "a") as log_file:
            log_file.write(error_message + "\n")
        return None