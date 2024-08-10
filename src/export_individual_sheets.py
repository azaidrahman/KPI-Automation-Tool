import pandas as pd
from tkinter import messagebox,filedialog
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def export_csm_sheet(workbook_path, csm_name):
    try:
        # Load the entire workbook
        wb = load_workbook(workbook_path)
        
        if csm_name in wb.sheetnames:
            # Create a new workbook for the exported sheet
            new_wb = load_workbook(filename=workbook_path)
            new_wb.remove(new_wb['Index'])  # Remove the Index sheet from the new workbook
            
            # Keep only the desired CSM sheet in the new workbook
            for sheet_name in new_wb.sheetnames:
                if sheet_name != csm_name:
                    new_wb.remove(new_wb[sheet_name])
            
            # Adjust column widths
            sheet = new_wb[csm_name]
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the new workbook
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=f"{csm_name}.xlsx")
            if save_path:
                new_wb.save(save_path)
                messagebox.showinfo("Success", f"Sheet for {csm_name} exported successfully.")
        else:
            messagebox.showerror("Error", f"Sheet for {csm_name} not found in the workbook.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while exporting: {str(e)}")

def export_all_sheets(workbook_path):
    try:
        # Load the entire workbook
        wb = load_workbook(workbook_path)
        export_folder = filedialog.askdirectory()
        if export_folder:
            for sheet_name in wb.sheetnames:
                if sheet_name != 'Index':
                    # Create a new workbook for each sheet
                    new_wb = load_workbook(filename=workbook_path)
                    new_wb.remove(new_wb['Index'])  # Remove the Index sheet
                    
                    # Keep only the current sheet in the new workbook
                    for name in new_wb.sheetnames:
                        if name != sheet_name:
                            new_wb.remove(new_wb[name])
                    
                    # Adjust column widths
                    sheet = new_wb[sheet_name]
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Save the new workbook
                    new_wb.save(os.path.join(export_folder, f"{sheet_name}.xlsx"))
            messagebox.showinfo("Success", "All sheets exported successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while exporting all sheets: {str(e)}")